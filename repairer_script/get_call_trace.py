import os
import openpyxl
import argparse
from collections import defaultdict

def write_statistics_to_excel(extracted_data, excel_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Bug Statistics'
    sheet.append(["Bug Name", "Reproducer Available", "Reproducer Type", "Bug Location"])
    
    for bug_name, reproducer_type, _, bug_location in extracted_data:
        reproducer_available = 'Yes' if reproducer_type != 'N/A' else 'No'
        sheet.append([bug_name, reproducer_available, reproducer_type, bug_location])

    workbook.save(excel_file)

def write_unique_traces_to_excel(trace_groups, excel_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Unique Call Traces'
    sheet.append(["Call Trace", "Bug Name"])
    
    for call_trace, bug_names in trace_groups.items():
        if len(bug_names) == 1:  # Unique call traces only
            sheet.append([call_trace, bug_names[0]])
    
    workbook.save(excel_file)

def group_by_call_traces(extracted_data):
    # Create a dictionary to group bug names by call trace
    trace_groups = defaultdict(list)
    for bug_name, reproducer_present, call_trace, bug_location in extracted_data:
        trace_groups[call_trace].append(bug_name)
    return trace_groups

def write_identical_traces_to_excel(trace_groups, excel_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Identical Call Traces'
    sheet.append(["Call Trace", "Bug Names", "Count"])
    
    for call_trace, bug_names in trace_groups.items():
        # Only consider call traces with more than one bug name
        if len(bug_names) > 1:
            sheet.append([call_trace, ', '.join(bug_names), len(bug_names)])
    
    workbook.save(excel_file)

def extract_call_trace(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
    except UnicodeDecodeError:
        try:
            with open(file_path, 'r', encoding='latin-1') as file:
                content = file.read()
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                content = file.read()

    start = content.find('<TASK>')
    end = content.find('</TASK>', start)
    if start != -1 and end != -1:
        # Extract the content without <TASK> and </TASK>
        call_trace_content = content[start + len('<TASK>'):end].strip()
        
        # Remove specific lines
        filtered_content = []
        for line in call_trace_content.split('\n'):
            if not any(line.startswith(prefix) for prefix in ['RIP:', 'Code:', 'RSP:', 'RAX:', 'RDX:', 'RBP:', 'R10:', 'R13:']):
                filtered_content.append(line)
        return '\n'.join(filtered_content)
    return None

def write_to_excel(extracted_data, excel_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Bug Reports'
    sheet.append(["Bug Name", "Call Trace"])
    
    # Set the column width to ensure visibility
    sheet.column_dimensions['B'].width = 100
    
    for bug_name, reproducer_present, call_trace, bug_location in extracted_data:
        # Append the data to the sheet
        sheet.append([bug_name, call_trace])
        # Get the last row and column number where data was inserted
        row = sheet.max_row
        col = 2  # Assuming call trace is in the second column
        cell = sheet.cell(row=row, column=col)
        
        # Set the alignment for the call trace to wrap text
        cell.alignment = openpyxl.styles.Alignment(wrapText=True)
        
        # Calculate the approximate height needed for the cell (one line is roughly 15 points high)
        number_of_lines = call_trace.count('\n') + 1
        cell_height = 15 * number_of_lines
        sheet.row_dimensions[row].height = cell_height  # Set the row height

    workbook.save(excel_file)

def extract_from_directory(base_dir):
    extracted_data = []
    exclude_list = ['BUG: MAX_LOCKDEP_ENTRIES too low!', 'suppressed report']
    
    for root, dirs, files in os.walk(base_dir):
        if 'description' in files:
            with open(os.path.join(root, 'description'), 'r') as desc_file:
                bug_name = desc_file.read().strip()

            reproducer_type = 'N/A'
            if 'repro.cprog' in files:
                reproducer_type = 'C reproducer'
            elif any(repro_file in files for repro_file in ['repro0', 'repro.prog']):
                reproducer_type = 'Syz reproducer'

            if bug_name not in exclude_list:
                bug_added = False
                for file in files:
                    if file in ['report0', 'repro.report'] and not bug_added:
                        call_trace = extract_call_trace(os.path.join(root, file))
                        if call_trace:
                            extracted_data.append((bug_name, reproducer_type, call_trace, root))
                            bug_added = True  # Set the flag to True after adding the bug
    return extracted_data

def parse_arguments():
    parser = argparse.ArgumentParser(description="Extract call traces from syzkaller bug reports.")
    parser.add_argument("--bug-directory", required=True, help="Directory containing the bug reports")
    parser.add_argument("--output", required=True, help="Output Excel file")
    parser.add_argument("--identical-traces", required=True, help="Output Identical Traces Excel file")
    parser.add_argument("--unique-traces", required=True, help="Output Unique Traces Excel file")
    parser.add_argument("--statistics", required=True, help="Output Statistics Excel file")
    return parser.parse_args()

if __name__ == '__main__':
    args = parse_arguments()
    extracted_data = extract_from_directory(args.bug_directory)
    
    if extracted_data:
        write_to_excel(extracted_data, args.output)

        trace_groups = group_by_call_traces(extracted_data)
        write_identical_traces_to_excel(trace_groups, args.identical_traces)
        write_unique_traces_to_excel(trace_groups, args.unique_traces)

        write_statistics_to_excel(extracted_data, args.statistics)
    else:
        print("No data extracted.")
